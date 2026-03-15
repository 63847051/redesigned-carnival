#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FinanceDatabase 报告生成器
支持多格式导出（CSV、JSON、PDF、Excel）
"""

import pandas as pd
import json
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not installed. PDF generation disabled.")


class ReportGenerator:
    """报告生成器基类"""

    def __init__(self, output_dir: str = "/tmp/financedatabase_reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def _get_output_path(self, filename: str, extension: str) -> Path:
        """获取输出文件路径"""
        return self.output_dir / f"{filename}_{self.timestamp}.{extension}"


class CSVReportGenerator(ReportGenerator):
    """CSV 报告生成器"""

    def generate(self, data: pd.DataFrame, filename: str = "report") -> str:
        """
        生成 CSV 报告
        """
        output_path = self._get_output_path(filename, "csv")
        data.to_csv(output_path, index=False, encoding='utf-8-sig')
        return str(output_path)

    def generate_multiple(self, data_dict: Dict[str, pd.DataFrame], filename: str = "reports") -> str:
        """
        生成多个 CSV 文件（打包到一个目录）
        """
        base_dir = self.output_dir / f"{filename}_{self.timestamp}"
        base_dir.mkdir(parents=True, exist_ok=True)

        for name, data in data_dict.items():
            output_path = base_dir / f"{name}.csv"
            data.to_csv(output_path, index=False, encoding='utf-8-sig')

        return str(base_dir)


class JSONReportGenerator(ReportGenerator):
    """JSON 报告生成器"""

    def generate(self, data: pd.DataFrame, filename: str = "report", orient: str = "records") -> str:
        """
        生成 JSON 报告
        """
        output_path = self._get_output_path(filename, "json")
        data.to_json(output_path, orient=orient, force_ascii=False, indent=2)
        return str(output_path)

    def generate_from_dict(self, data: Dict, filename: str = "report") -> str:
        """
        从字典生成 JSON 报告
        """
        output_path = self._get_output_path(filename, "json")
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return str(output_path)


class ExcelReportGenerator(ReportGenerator):
    """Excel 报告生成器"""

    def generate(self, data: pd.DataFrame, filename: str = "report", sheet_name: str = "Sheet1") -> str:
        """
        生成 Excel 报告
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation. Install: pip install openpyxl")

        output_path = self._get_output_path(filename, "xlsx")
        data.to_excel(output_path, sheet_name=sheet_name, index=False, engine='openpyxl')
        return str(output_path)

    def generate_multiple_sheets(self, data_dict: Dict[str, pd.DataFrame], filename: str = "report") -> str:
        """
        生成多 Sheet Excel 文件
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation. Install: pip install openpyxl")

        output_path = self._get_output_path(filename, "xlsx")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, data in data_dict.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        return str(output_path)

    def generate_with_formatting(self, data: pd.DataFrame, filename: str = "report") -> str:
        """
        生成带格式的 Excel 报告（添加样式）
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required. Install: pip install openpyxl")

        output_path = self._get_output_path(filename, "xlsx")

        # 先写入数据
        data.to_excel(output_path, index=False, engine='openpyxl')

        # 添加格式
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # 标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 数据行样式
        cell_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用标题样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 应用数据样式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell_alignment
                cell.border = border

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        return str(output_path)


class PDFReportGenerator(ReportGenerator):
    """PDF 报告生成器"""

    def __init__(self, output_dir: str = "/tmp/financedatabase_reports"):
        super().__init__(output_dir)
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF generation")

    def generate_table_report(self, data: pd.DataFrame, filename: str = "report",
                             title: str = "Financial Report") -> str:
        """
        生成表格 PDF 报告
        """
        output_path = self._get_output_path(filename, "pdf")

        doc = SimpleDocTemplate(str(output_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # 标题
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.darkblue,
            spaceAfter=30
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))

        # 生成时间
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        story.append(Paragraph(f"<b>生成时间:</b> {timestamp}", styles['Normal']))
        story.append(Spacer(1, 12))

        # 数据统计
        story.append(Paragraph(f"<b>数据行数:</b> {len(data)}", styles['Normal']))
        story.append(Paragraph(f"<b>数据列数:</b> {len(data.columns)}", styles['Normal']))
        story.append(Spacer(1, 12))

        # 转换数据为列表
        data_list = [data.columns.tolist()] + data.values.tolist()

        # 创建表格
        table = Table(data_list)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(table)
        doc.build(story)

        return str(output_path)

    def generate_health_report(self, health_data: Dict, filename: str = "health_report") -> str:
        """
        生成健康度分析 PDF 报告
        """
        output_path = self._get_output_path(filename, "pdf")

        doc = SimpleDocTemplate(str(output_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # 标题
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.darkblue,
            spaceAfter=20,
            alignment=1  # Center
        )
        story.append(Paragraph("财务健康度分析报告", title_style))
        story.append(Spacer(1, 12))

        # 生成时间
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        story.append(Paragraph(f"<b>生成时间:</b> {timestamp}", styles['Normal']))
        story.append(Spacer(1, 12))

        # 总体评分
        overall_score = health_data.get('scores', {}).get('overall', 'N/A')
        overall_rating = health_data.get('overall_rating', 'N/A')

        story.append(Paragraph(f"<b>总体评分:</b> {overall_score}", styles['Heading2']))
        story.append(Paragraph(f"<b>总体评级:</b> {overall_rating}", styles['Heading2']))
        story.append(Spacer(1, 12))

        # 各分类评分
        story.append(Paragraph("<b>分类评分详情:</b>", styles['Heading2']))

        table_data = [['分类', '评分', '评级', '建议']]
        for category, info in health_data.get('scores', {}).items():
            if category == 'overall':
                continue
            table_data.append([
                info.get('name', category),
                str(info.get('score', 'N/A')),
                info.get('rating', 'N/A'),
                info.get('advice', '')
            ])

        table = Table(table_data, colWidths=[2*inch, 1*inch, 1*inch, 2.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('WORDWRAP', (0, 0), (-1, -1), True)
        ]))

        story.append(table)
        story.append(Spacer(1, 12))

        # 改进建议
        if health_data.get('recommendations'):
            story.append(Paragraph("<b>改进建议:</b>", styles['Heading2']))
            for rec in health_data['recommendations']:
                story.append(
                    Paragraph(
                        f"• <b>{rec['category']} (优先级: {rec['priority']})</b>: {rec['action']}",
                        styles['Normal']
                    )
                )

        doc.build(story)

        return str(output_path)


class ComprehensiveReportGenerator:
    """综合报告生成器"""

    def __init__(self, output_dir: str = "/tmp/financedatabase_reports"):
        self.output_dir = output_dir
        self.csv_gen = CSVReportGenerator(output_dir)
        self.json_gen = JSONReportGenerator(output_dir)
        self.excel_gen = ExcelReportGenerator(output_dir)

        if REPORTLAB_AVAILABLE:
            self.pdf_gen = PDFReportGenerator(output_dir)

    def generate_all_formats(self, data: pd.DataFrame, filename: str = "report") -> Dict[str, str]:
        """
        生成所有格式的报告
        """
        results = {
            'csv': self.csv_gen.generate(data, filename),
            'json': self.json_gen.generate(data, filename),
            'excel': self.excel_gen.generate_with_formatting(data, filename)
        }

        if REPORTLAB_AVAILABLE:
            results['pdf'] = self.pdf_gen.generate_table_report(data, filename)

        return results

    def generate_dashboard_report(self, health_data: Dict, stock_data: pd.DataFrame,
                                  filename: str = "dashboard") -> Dict[str, str]:
        """
        生成仪表板报告（健康度 + 股票数据）
        """
        results = {}

        # 健康度报告
        if REPORTLAB_AVAILABLE:
            results['health_pdf'] = self.pdf_gen.generate_health_report(health_data, f"{filename}_health")

        # 股票数据报告
        results['stocks_excel'] = self.excel_gen.generate_with_formatting(stock_data, f"{filename}_stocks")
        results['stocks_csv'] = self.csv_gen.generate(stock_data, f"{filename}_stocks")

        # 综合数据
        combined_data = {
            'health_report': health_data,
            'stock_data': stock_data.to_dict(orient='records'),
            'generated_at': datetime.now().isoformat()
        }
        results['combined_json'] = self.json_gen.generate_from_dict(combined_data, filename)

        return results


def demo_usage():
    """演示使用"""
    print("=== 报告生成器演示 ===\n")

    # 创建示例数据
    data = pd.DataFrame({
        'symbol': ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA'],
        'price': [175.5, 378.2, 141.8, 178.3, 242.6],
        'pe_ratio': [25.5, 28.3, 22.1, 45.6, 85.3],
        'market_cap': [2500, 2300, 1800, 1500, 800]
    })

    # 健康度数据
    health_data = {
        'scores': {
            'liquidity': {'score': 85.5, 'name': '流动性健康', 'rating': '良好', 'advice': '保持良好现金流'},
            'profitability': {'score': 92.3, 'name': '盈利能力', 'rating': '优秀', 'advice': '盈利能力极强'},
            'solvency': {'score': 78.2, 'name': '偿债能力', 'rating': '良好', 'advice': '负债水平合理'},
            'efficiency': {'score': 88.9, 'name': '运营效率', 'rating': '良好', 'advice': '资产利用率高'},
            'overall': 86.23
        },
        'overall_rating': '良好',
        'overall_advice': '财务状况健康，可适当优化',
        'recommendations': [
            {'category': '偿债能力', 'priority': '中', 'action': '降低负债率至 60% 以下'}
        ]
    }

    # 生成报告
    generator = ComprehensiveReportGenerator()

    print("1. 生成 CSV 报告...")
    csv_path = generator.csv_gen.generate(data, "demo")
    print(f"   ✓ CSV: {csv_path}")

    print("\n2. 生成 JSON 报告...")
    json_path = generator.json_gen.generate(data, "demo")
    print(f"   ✓ JSON: {json_path}")

    print("\n3. 生成 Excel 报告...")
    excel_path = generator.excel_gen.generate_with_formatting(data, "demo")
    print(f"   ✓ Excel: {excel_path}")

    if REPORTLAB_AVAILABLE:
        print("\n4. 生成 PDF 报告...")
        pdf_path = generator.pdf_gen.generate_table_report(data, "demo")
        print(f"   ✓ PDF: {pdf_path}")

        print("\n5. 生成健康度 PDF 报告...")
        health_pdf_path = generator.pdf_gen.generate_health_report(health_data)
        print(f"   ✓ 健康度 PDF: {health_pdf_path}")

    print("\n6. 生成所有格式报告...")
    all_reports = generator.generate_all_formats(data, "comprehensive")
    for format_type, path in all_reports.items():
        print(f"   ✓ {format_type.upper()}: {path}")


if __name__ == "__main__":
    demo_usage()
